# -*- coding: utf-8 -*-
"""Command-line interface for ASIN PricePulse 价格脉搏.

Usage:
    pricepulse --markets DE,UK,US --input asins.csv --output ./out
    pricepulse --markets DE --asins B0DCBB2YTR,B09B96TG33 --output ./out

Any agent (Orcha, Claude Code, Kiro, Cursor) — or a human — can invoke this
command. Input formats: CSV, TXT, or comma-separated ASIN list on CLI.
"""
from __future__ import annotations

import argparse
import csv
import sys
import time
from pathlib import Path

from . import __version__
from .fetcher import MARKETS, fetch_batch, PriceResult
from .reporter import write_csv, write_html


import re as _re


def _normalize_header(h: str) -> str:
    """Lowercase + strip emojis/whitespace/punctuation for header matching."""
    if not h:
        return ""
    s = str(h).lower()
    # Remove emojis and non-ASCII symbols
    s = _re.sub(r'[^\w\s]', ' ', s)
    s = _re.sub(r'\s+', ' ', s).strip()
    return s


def _find_col(headers: list[str], keywords: list[str]) -> int:
    """Find column index whose normalized header contains ALL given keywords."""
    for i, h in enumerate(headers):
        norm = _normalize_header(h)
        if all(k in norm for k in keywords):
            return i
    return -1


def _to_number(v):
    """Convert to float if possible, else None."""
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def _read_input(path: str) -> list[dict]:
    """Read ASINs (and optional Search Rank / Purchase Rank) from file.

    Returns list of dicts: [{"asin": "...", "search_rank": float|None,
                              "purchase_rank": float|None}, ...]

    Supported: .xlsx, .csv, .tsv, plain text.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    rows: list[list] = []

    # ---------- xlsx ----------
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError(
                "Reading .xlsx requires openpyxl. Run: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

    # ---------- csv / tsv ----------
    elif p.suffix.lower() in (".csv", ".tsv"):
        delim = "\t" if p.suffix.lower() == ".tsv" else ","
        with open(p, "r", encoding="utf-8-sig", newline="") as f:
            rows = [list(r) for r in csv.reader(f, delimiter=delim)]

    # ---------- plain text ----------
    else:
        text = p.read_text(encoding="utf-8-sig")
        items = []
        for line in text.splitlines():
            for cell in line.replace(",", " ").replace("\t", " ").split():
                cell = cell.strip()
                if cell:
                    items.append({"asin": cell})
        return items

    if not rows:
        return []

    # Detect header row
    first_row = rows[0]
    start = 0
    # Header exists if first row doesn't have a 10-char ASIN in any cell
    has_asin_in_first_row = any(
        isinstance(c, str) and len(c.strip()) == 10 and c.strip().isalnum()
        for c in first_row if c is not None
    )
    if not has_asin_in_first_row:
        start = 1
        headers = [str(c) if c is not None else "" for c in first_row]
    else:
        headers = []

    # Find columns
    volume_cols_map: dict[int, str] = {}  # column-index → date label (for time series)
    if headers:
        asin_col = _find_col(headers, ["asin"])
        search_col = _find_col(headers, ["search", "rank"])
        purchase_col = _find_col(headers, ["purchase", "rank"])
        # "Keyword Searches", "Search Volume", "Searches", "Volume"
        volume_col = -1
        for kws in [["keyword", "search"], ["search", "volume"],
                    ["searches"], ["volume"]]:
            volume_col = _find_col(headers, kws)
            if volume_col >= 0 and volume_col != search_col:
                break
        # Time-series columns: header matches YYYY-MM / YYYY/MM / MMM YYYY etc.
        _DATE_RE = _re.compile(
            r'^\s*('
            r'\d{4}[-/]\d{1,2}'                   # 2023-01, 2023/1
            r'|(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s*\d{2,4}'
            r'|\d{1,2}[-/]\d{4}'                  # 01-2023
            r')\s*$',
            _re.IGNORECASE,
        )
        for i, h in enumerate(headers):
            if _DATE_RE.match(str(h) if h is not None else ""):
                volume_cols_map[i] = str(h).strip()
    else:
        asin_col = 0
        search_col = -1
        purchase_col = -1
        volume_col = -1

    if asin_col < 0:
        asin_col = 0  # fallback: first column

    items = []
    for r in rows[start:]:
        if asin_col >= len(r):
            continue
        v = r[asin_col]
        if v is None:
            continue
        asin = str(v).strip().upper()
        if not asin or not (len(asin) == 10 and asin.isalnum()):
            continue
        item = {"asin": asin}
        if search_col >= 0 and search_col < len(r):
            item["search_rank"] = _to_number(r[search_col])
        if purchase_col >= 0 and purchase_col < len(r):
            item["purchase_rank"] = _to_number(r[purchase_col])
        if volume_col >= 0 and volume_col < len(r):
            item["search_volume"] = _to_number(r[volume_col])
        if volume_cols_map:
            series = {}
            for ci, label in volume_cols_map.items():
                if ci < len(r):
                    val = _to_number(r[ci])
                    if val is not None:
                        series[label] = val
            if series:
                item["volume_series"] = series
        items.append(item)
    return items


def _estimate_runtime(total_requests: int) -> str:
    """Rough estimate accounting for Amazon rate-limiting on large batches."""
    if total_requests <= 30:
        return "~1-2 min"
    if total_requests <= 100:
        return "~3-8 min"
    if total_requests <= 300:
        return "~15-40 min (Amazon rate-limits after ~50 requests)"
    return f"~{total_requests // 6} min or more (consider batching)"


def _print_pre_flight(markets: list[str], items: list, extras: list[str] = None) -> None:
    total_req = len(items) * len(markets)
    print("=" * 70, file=sys.stderr)
    print(" 🔍 ASIN PricePulse 价格脉搏 v" + __version__ + " · pre-flight check", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print(f" Markets: {', '.join(markets)}", file=sys.stderr)
    print(f" ASINs:   {len(items)} (deduped)", file=sys.stderr)
    print(f" Total requests: {total_req}", file=sys.stderr)
    print(f" Estimated runtime: {_estimate_runtime(total_req)}", file=sys.stderr)
    if extras:
        print(f" Extra columns detected: {', '.join(extras)}", file=sys.stderr)
    if total_req > 100:
        print(" ⚠  Large batch — Amazon will progressively rate-limit.", file=sys.stderr)
        print("    For >300 requests, consider --delay 1.0 --workers 3 to smooth pace.", file=sys.stderr)
    print("", file=sys.stderr)
    print(" i Reminders:", file=sys.stderr)
    print("   • Prices reflect what a local guest shopper sees at this moment.", file=sys.stderr)
    print("   • Delivery address is auto-set per market (e.g. Berlin for DE).", file=sys.stderr)
    print("   • Currency is forced to the market's native currency.", file=sys.stderr)
    print("   • Your VPN state does NOT affect results (cookie-based routing).", file=sys.stderr)
    print("   • For AST-internal use only. Do not share externally without", file=sys.stderr)
    print("     stripping internal attribution and disclaimers.", file=sys.stderr)
    print("=" * 70, file=sys.stderr)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        prog="pricepulse",
        description="ASIN PricePulse 价格脉搏 — Batch-fetch Amazon ASIN prices across 17 marketplaces.",
    )
    ap.add_argument("--markets", "-m", required=True,
                    help="Comma-separated market codes (e.g. US,DE,UK). "
                         f"Available: {','.join(sorted(MARKETS.keys()))}")
    inp = ap.add_mutually_exclusive_group(required=True)
    inp.add_argument("--input", "-i", help="Path to CSV/TXT of ASINs")
    inp.add_argument("--asins", "-a",
                     help="Comma-separated ASINs (e.g. B0DCBB2YTR,B09B96TG33)")
    ap.add_argument("--output", "-o", default=".",
                    help="Output directory (default: current dir)")
    ap.add_argument("--workers", "-w", type=int, default=6,
                    help="Parallel workers per market (default 6)")
    ap.add_argument("--delay", type=float, default=0.4,
                    help="Delay between requests, seconds (default 0.4)")
    ap.add_argument("--title", default="ASIN Price Intelligence Report",
                    help="HTML report title")
    ap.add_argument("--subtitle", default="",
                    help="Optional subtitle below title, e.g. date range '2026.5.1 - 7.31'")
    ap.add_argument("--yes", "-y", action="store_true",
                    help="Skip pre-flight confirmation prompt")
    ap.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    args = ap.parse_args(argv)

    markets = [m.strip().upper() for m in args.markets.split(",") if m.strip()]
    if args.input:
        items = _read_input(args.input)
    else:
        items = [{"asin": a.strip().upper()} for a in args.asins.split(",") if a.strip()]

    if not items:
        print("ERROR: no ASINs provided.", file=sys.stderr)
        return 2
    if not markets:
        print("ERROR: no markets provided.", file=sys.stderr)
        return 2

    has_search_rank = any(it.get("search_rank") is not None for it in items)
    has_purchase_rank = any(it.get("purchase_rank") is not None for it in items)
    has_search_volume = any(it.get("search_volume") is not None for it in items)
    has_volume_series = any(it.get("volume_series") for it in items)
    extras = []
    if has_search_rank: extras.append("Search Rank")
    if has_purchase_rank: extras.append("Purchase Rank")
    if has_search_volume: extras.append("Search Volume")
    if has_volume_series: extras.append("Volume Time-Series")

    _print_pre_flight(markets, items, extras)

    if not args.yes and sys.stdin.isatty():
        confirm = input(" Proceed? [Y/n] ").strip().lower()
        if confirm and confirm not in ("y", "yes"):
            print("Aborted.", file=sys.stderr)
            return 1

    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")

    t0 = time.time()
    last_print = [0.0]

    def progress(done, total, market, asin):
        now = time.time()
        if now - last_print[0] > 0.3 or done == total:
            pct = done / total * 100
            print(f"  [{done}/{total}] {pct:5.1f}%  {market}  {asin}",
                  file=sys.stderr)
            last_print[0] = now

    results = fetch_batch(
        asins=items, markets=markets,
        max_workers=args.workers, delay_between=args.delay,
        progress_callback=progress,
    )
    elapsed = time.time() - t0

    detail_path = out_dir / f"asin_detail_{ts}.xlsx"
    html_path = out_dir / f"\u7ade\u54c1ASIN\u5206\u6790\u62a5\u544a_{ts}.html"
    write_csv(results, detail_path)
    write_html(results, html_path, title=args.title, subtitle=args.subtitle)

    ok = sum(1 for r in results if r.status == "ok")
    print("", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print(f" ✓ Done in {elapsed:.1f}s · {ok}/{len(results)} priced successfully",
          file=sys.stderr)
    print(f" ✓ Excel: {detail_path.resolve()}", file=sys.stderr)
    print(f" ✓ HTML:  {html_path.resolve()}", file=sys.stderr)
    print("=" * 70, file=sys.stderr)

    # Print machine-readable summary to stdout for agents
    print(str(html_path.resolve()))
    return 0


if __name__ == "__main__":
    sys.exit(main())
