# -*- coding: utf-8 -*-
"""ASIN PriceLens reporter — generates Excel and interactive HTML reports.

v1.2.0 — 2026-08-11. New layout: price tier cards + vertical bar chart + brand avg price chart.
"""
from __future__ import annotations

import json
import html as html_mod
from pathlib import Path
from dataclasses import asdict
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .fetcher import PriceResult


# ===========================================================================
# Excel output
# ===========================================================================

def write_csv(results: list, path) -> None:
    """Write results to styled Excel file (.xlsx)."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        path = path.with_suffix(".xlsx")

    if not results:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.append(["No results"])
        wb.save(path)
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    fieldnames = [
        "asin", "market", "currency", "price", "display_price",
        "title", "brand", "availability", "status", "url", "error",
        "search_rank", "purchase_rank", "search_volume",
    ]
    header_names = [
        "ASIN", "Market", "Currency", "Price", "Display Price",
        "Title", "Brand", "Availability", "Status", "URL", "Error",
        "Search Rank", "Purchase Rank", "Search Volume",
    ]
    series_keys: list[str] = []
    for r in results:
        d = asdict(r) if not isinstance(r, dict) else r
        vs = d.get("volume_series")
        if vs and isinstance(vs, dict):
            for k in vs:
                if k not in series_keys:
                    series_keys.append(k)
    series_keys.sort()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASIN Detail"

    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Segoe UI", size=10)
    body_align = Alignment(vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = header_names + series_keys
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r in results:
        d = asdict(r) if not isinstance(r, dict) else r
        row = [d.get(fn, "") for fn in fieldnames]
        vs = d.get("volume_series") or {}
        for sk in series_keys:
            row.append(vs.get(sk, ""))
        ws.append(row)

    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    for col_idx in range(1, ws.max_column + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(ws.max_row + 1, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    ws.freeze_panes = "A2"
    wb.save(path)


# ===========================================================================
# HTML output
# ===========================================================================

_TIER_COLORS = ["#16A34A", "#2563EB", "#DC2626"]  # green, blue, red for Entry/Mid/Premium
_BAR_COLORS = ["#7C3AED", "#F59E0B", "#16A34A", "#DC2626", "#2563EB",
               "#0891B2", "#EA580C", "#4F46E5", "#059669", "#DB2777",
               "#7C3AED", "#F59E0B"]


def write_html(results: list, path, title: str = "ASIN Price Intelligence Report",
               subtitle: str = "") -> None:
    """Generate interactive HTML report."""
    path = Path(path)
    if not results:
        path.write_text("<!DOCTYPE html><html><body><p>No results</p></body></html>",
                        encoding="utf-8")
        return

    rows = [asdict(r) if not isinstance(r, dict) else r for r in results]

    has_search_rank = any(r.get("search_rank") is not None for r in rows)
    has_purchase_rank = any(r.get("purchase_rank") is not None for r in rows)
    has_search_volume = any(r.get("search_volume") is not None for r in rows)
    has_volume_series = any(r.get("volume_series") for r in rows)

    markets_data: dict[str, list[dict]] = {}
    for r in rows:
        mk = r.get("market", "Unknown")
        markets_data.setdefault(mk, []).append(r)

    panels_html = ""
    tabs_html = ""
    first = True
    for mk, mrows in markets_data.items():
        active = " active" if first else ""
        tabs_html += f'<div class="tab{active}" data-market="{mk}">{mk} ({len(mrows)})</div>\n'
        panels_html += _render_market_panel(mk, mrows, active, has_search_rank,
                                            has_purchase_rank, has_search_volume,
                                            has_volume_series)
        first = False

    json_rows = [{k: v for k, v in r.items() if k != "volume_series"} for r in rows]
    volume_series_json = {}
    if has_volume_series:
        for r in rows:
            vs = r.get("volume_series")
            if vs:
                volume_series_json[r["asin"]] = vs

    escaped_title = html_mod.escape(title)
    escaped_subtitle = html_mod.escape(subtitle) if subtitle else ""
    html_content = _render_html(escaped_title, escaped_subtitle, tabs_html, panels_html,
                                json_rows, volume_series_json, has_search_rank,
                                has_purchase_rank, has_search_volume, has_volume_series)
    path.write_text(html_content, encoding="utf-8")


def _price_bands(prices: list[float]) -> list[tuple[str, int]]:
    """Equal-width 3-tier price banding."""
    if not prices:
        return []
    sorted_p = sorted(prices)
    mn, mx = sorted_p[0], sorted_p[-1]
    step = (mx - mn) / 3 if mx > mn else max(mx * 0.1, 1)
    bounds = [mn, mn + step, mn + step * 2, mx]
    labels = ["Entry", "Mid-tier", "Premium"]
    bands = []
    for i in range(3):
        lo, hi = bounds[i], bounds[i + 1]
        if i == 2:
            count = sum(1 for p in prices if lo <= p <= hi)
        else:
            count = sum(1 for p in prices if lo <= p < hi)
        bands.append((f"{labels[i]}\n({lo:.0f}\u2013{hi:.0f})" if i < 2
                      else f"{labels[i]}\n({lo:.0f}+)", count))
    return bands


def _summarize_market(mrows: list[dict]) -> dict:
    ok_rows = [r for r in mrows if r.get("status") == "ok" and r.get("price")]
    prices = [r["price"] for r in ok_rows]
    brands = set(r.get("brand", "") for r in ok_rows if r.get("brand"))
    return {
        "total": len(mrows), "ok": len(ok_rows),
        "min_price": min(prices) if prices else 0,
        "max_price": max(prices) if prices else 0,
        "avg_price": sum(prices) / len(prices) if prices else 0,
        "median_price": sorted(prices)[len(prices) // 2] if prices else 0,
        "brands_count": len(brands),
        "currency": ok_rows[0]["currency"] if ok_rows else "",
    }


def _render_market_panel(mk: str, mrows: list[dict], active: str,
                         has_sr: bool, has_pr: bool, has_sv: bool,
                         has_vs: bool) -> str:
    summary = _summarize_market(mrows)
    cur = summary["currency"]
    display = "block" if active else "none"

    ok_rows = [r for r in mrows if r.get("status") == "ok" and r.get("price")]
    prices = [r["price"] for r in ok_rows]
    bands = _price_bands(prices)

    # --- Price Tier Cards (top section) ---
    tier_cards_html = ""
    if bands:
        tier_labels = ["Entry", "Mid-tier", "Premium"]
        tier_colors = _TIER_COLORS
        total_priced = sum(c for _, c in bands)
        for i, (label, count) in enumerate(bands):
            parts = label.split("\n")
            tier_name = parts[0]
            tier_range = parts[1] if len(parts) > 1 else ""
            pct = f"{count/total_priced*100:.1f}" if total_priced else "0"
            # Compute avg price and avg search rank for this tier
            tier_prices = []
            tier_ranks = []
            # Determine bounds for this tier
            mn_p = min(prices) if prices else 0
            mx_p = max(prices) if prices else 0
            step = (mx_p - mn_p) / 3 if mx_p > mn_p else 1
            lo = mn_p + step * i
            hi = mn_p + step * (i + 1)
            for r in ok_rows:
                p = r["price"]
                if i == 2:
                    in_tier = p >= lo
                else:
                    in_tier = lo <= p < hi
                if in_tier:
                    tier_prices.append(p)
                    sr = r.get("search_rank")
                    if sr is not None:
                        tier_ranks.append(sr)
            avg_p = sum(tier_prices) / len(tier_prices) if tier_prices else 0
            avg_rank = sum(tier_ranks) / len(tier_ranks) if tier_ranks else 0
            # Representative ASIN (highest search volume in tier)
            rep_asin = ""
            if has_sv:
                tier_vol = [(r.get("asin", ""), r.get("search_volume") or 0) for r in ok_rows
                            if (r["price"] >= lo and (r["price"] < hi if i < 2 else True))]
                tier_vol.sort(key=lambda x: -x[1])
                rep_asin = tier_vol[0][0] if tier_vol else ""

            tier_cards_html += f'''
            <div class="tier-card clickable-bar" data-filter-type="price" data-filter-value="{tier_name}" data-market="{mk}" style="border-top:4px solid {tier_colors[i]}">
              <div class="tier-header">{tier_name} \u00b7 ${tier_range.replace("\u2013"," \u2013 $").replace("+","+")}
</div>
              <div class="tier-count">{count} <span class="tier-pct">({pct}%)</span></div>
              <div class="tier-meta"><span class="tier-meta-label">Avg Price</span><span class="tier-meta-val">${avg_p:.2f}</span></div>
              <div class="tier-meta"><span class="tier-meta-label">Avg Search Rank</span><span class="tier-meta-val">#{avg_rank:.0f}</span></div>
              {f'<div class="tier-meta"><span class="tier-meta-label">Top ASIN</span><span class="tier-meta-val tier-asin">{rep_asin}</span></div>' if rep_asin else ''}
            </div>'''

    # --- Brand bubble cloud ---
    brand_counts: dict[str, int] = {}
    for r in ok_rows:
        b = r.get("brand", "").strip() or "Unknown"
        brand_counts[b] = brand_counts.get(b, 0) + 1
    top_brands_all = sorted(brand_counts.items(), key=lambda x: -x[1])
    # Bubble cloud (all brands with count)
    bubble_html = '<div class="section"><div class="section-title">Brand Distribution</div><div class="card"><div class="bubble-cloud">'
    for i, (bname, bcount) in enumerate(top_brands_all[:20]):
        color = _BAR_COLORS[i % len(_BAR_COLORS)]
        bubble_html += f'<span class="bubble clickable-bar" data-filter-type="brand" data-filter-value="{html_mod.escape(bname)}" data-market="{mk}" style="background:{color}">{html_mod.escape(bname)} <b>{bcount}</b></span>'
    bubble_html += '</div></div></div>'

    # --- Brand avg price chart (horizontal bar) ---
    brand_avg: list[tuple[str, float]] = []
    for bname, _ in top_brands_all[:12]:
        bp = [r["price"] for r in ok_rows if r.get("brand", "").strip() == bname]
        if bp:
            brand_avg.append((bname, sum(bp) / len(bp)))
    brand_avg.sort(key=lambda x: -x[1])
    brand_chart_html = ""
    if brand_avg:
        max_avg = brand_avg[0][1] or 1
        brand_chart_html = '<div class="card"><div class="card-title">Brand Avg Price \u00b7 TOP 12</div>'
        for i, (bname, avg) in enumerate(brand_avg):
            pct = avg / max_avg * 100
            color = _BAR_COLORS[i % len(_BAR_COLORS)]
            brand_chart_html += f'''<div class="hbar-row clickable-bar" data-filter-type="brand" data-filter-value="{html_mod.escape(bname)}" data-market="{mk}">
              <div class="hbar-label">{html_mod.escape(bname)}</div>
              <div class="hbar-track"><div class="hbar-fill" style="width:{pct:.1f}%;background:{color}"></div><span class="hbar-inline">${avg:.1f}</span></div>
            </div>'''
        brand_chart_html += '</div>'

    # --- Search Volume chart (JS-driven) ---
    vol_bar_html = ""
    if has_sv:
        vol_bar_html = f'<div class="section"><div class="section-title">Search Volume by ASIN (Top 15) <span class="vol-filter-label" id="vol-filter-{mk}"></span></div><div class="card"><div id="vol-chart-{mk}"></div></div></div>'

    # --- Detail table ---
    table_html = _render_detail_table(mk, mrows, has_sr, has_pr)

    return f'''
    <div class="tab-content{" active" if active else ""}" data-market="{mk}" style="display:{display}">
      <div class="section">
        <div class="section-title">\U0001f3c6 Price Tier Overview</div>
        <div class="tier-row">{tier_cards_html}</div>
      </div>
      <div class="section">
        <div class="grid-2">
          <div class="card"><div class="card-title">ASIN Count by Price Tier</div><canvas id="tier-bar-{mk}" height="220"></canvas></div>
          {brand_chart_html}
        </div>
      </div>
      {bubble_html}
      {vol_bar_html}
      <div class="section">
        <div class="section-title">ASIN Detail</div>
        {table_html}
      </div>
    </div>'''


def _render_detail_table(mk: str, mrows: list[dict],
                         has_sr: bool, has_pr: bool) -> str:
    cols = ["ASIN", "Brand", "Title", "Price", "Status"]
    if has_sr:
        cols.append("Search Rank")
    if has_pr:
        cols.append("Purchase Rank")

    th_html = ""
    for i, c in enumerate(cols):
        th_html += f'<th data-col="{i}">{html_mod.escape(c)} <span class="th-sort" data-col="{i}">&#8597;</span> <span class="th-filter" data-col="{i}">&#9662;</span></th>'

    tbody_rows = ""
    for r in mrows:
        price_str = f'{r.get("currency","")} {r["price"]:.2f}' if r.get("price") else "-"
        status = r.get("status", "unknown")
        title_short = (r.get("title", "") or "")[:80]
        cells = [
            f'<a href="{html_mod.escape(r.get("url",""))}" target="_blank">{html_mod.escape(r.get("asin",""))}</a>',
            html_mod.escape(r.get("brand", "") or ""),
            f'<span class="dim">{html_mod.escape(title_short)}</span>',
            price_str,
            f'<span class="st-{status}">{status}</span>',
        ]
        if has_sr:
            sv = r.get("search_rank")
            cells.append(f"{sv:.0f}" if sv is not None else "-")
        if has_pr:
            pv = r.get("purchase_rank")
            cells.append(f"{pv:.0f}" if pv is not None else "-")
        td_html = "".join(f"<td>{c}</td>" for c in cells)
        tbody_rows += f'<tr>{td_html}</tr>\n'

    return f'''
      <div class="table-scroll">
        <table id="table-{mk}">
          <thead><tr>{th_html}</tr></thead>
          <tbody>{tbody_rows}</tbody>
        </table>
      </div>'''


def _render_html(title: str, subtitle: str, tabs_html: str, panels_html: str,
                 json_rows: list[dict], volume_series_json: dict,
                 has_sr: bool, has_pr: bool, has_sv: bool, has_vs: bool) -> str:
    row_json = json.dumps(json_rows, ensure_ascii=False, separators=(",", ":"))
    vol_json = json.dumps(volume_series_json, ensure_ascii=False, separators=(",", ":"))

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#F8FAFC;color:#1E293B;font-size:13px;}}
a{{color:#2563EB;text-decoration:none;}}
a:hover{{text-decoration:underline;}}
.header{{background:linear-gradient(135deg,#1E3A5F 0%,#2563EB 100%);color:#fff;padding:28px 40px;text-align:center;}}
.header h1{{font-size:22px;font-weight:700;letter-spacing:.3px;}}
.header .sub{{font-size:13px;opacity:.85;margin-top:4px;}}
.container{{max-width:1400px;margin:0 auto;padding:24px 32px;}}
.section{{margin-bottom:28px;}}
.section-title{{font-size:15px;font-weight:700;color:#1E3A5F;border-left:4px solid #2563EB;padding-left:10px;margin-bottom:16px;}}
.tabs{{display:flex;gap:4px;margin-bottom:0;border-bottom:2px solid #E2E8F0;}}
.tab{{padding:7px 16px;font-size:12px;font-weight:600;cursor:pointer;border-radius:6px 6px 0 0;color:#64748B;border:1px solid transparent;border-bottom:none;position:relative;bottom:-2px;transition:all .15s;}}
.tab.active{{background:#fff;color:#2563EB;border-color:#E2E8F0;border-bottom-color:#fff;}}
.tab:hover:not(.active){{background:#F8FAFC;color:#1E293B;}}
.tab-content{{display:none;}}
.tab-content.active{{display:block;padding-top:16px;}}

/* Tier Cards */
.tier-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:16px;margin-bottom:20px;}}
.tier-card{{background:#fff;border-radius:12px;padding:20px 24px;box-shadow:0 1px 4px rgba(0,0,0,.07);cursor:pointer;transition:transform .15s,box-shadow .15s;}}
.tier-card:hover{{transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,0,0,.12);}}
.tier-card.selected{{box-shadow:0 0 0 2px #2563EB,0 4px 12px rgba(37,99,235,.2);}}
.tier-header{{font-size:12px;color:#64748B;margin-bottom:8px;}}
.tier-count{{font-size:26px;font-weight:800;color:#1E293B;}}
.tier-pct{{font-size:14px;font-weight:400;color:#94A3B8;}}
.tier-meta{{display:flex;justify-content:space-between;margin-top:8px;font-size:12px;}}
.tier-meta-label{{color:#64748B;}}
.tier-meta-val{{color:#1E293B;font-weight:600;}}
.tier-asin{{font-family:'SF Mono','Consolas',monospace;font-size:11px;}}

/* Grid */
.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:20px;}}
.card{{background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.07);}}
.card-title{{font-size:12px;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:.5px;margin-bottom:14px;}}

/* Bubble cloud */
.bubble-cloud{{display:flex;flex-wrap:wrap;gap:8px;}}
.bubble{{display:inline-flex;align-items:center;gap:6px;padding:6px 14px;border-radius:20px;color:#fff;font-size:12px;font-weight:500;cursor:pointer;transition:transform .15s,opacity .15s;}}
.bubble:hover{{transform:scale(1.05);}}
.bubble.selected{{outline:3px solid #1E293B;outline-offset:2px;}}
.bubble b{{font-weight:700;opacity:.9;}}

/* Horizontal bars */
.hbar-row{{display:grid;grid-template-columns:80px 1fr;align-items:center;gap:10px;margin-bottom:8px;cursor:pointer;border-radius:6px;padding:3px 6px;transition:background .15s;}}
.hbar-row:hover{{background:#F1F5F9;}}
.hbar-row.selected{{background:#EFF6FF;}}
.hbar-label{{font-size:12px;color:#334155;text-align:right;white-space:nowrap;}}
.hbar-track{{background:#F1F5F9;border-radius:4px;height:22px;display:flex;align-items:center;overflow:visible;}}
.hbar-fill{{height:100%;border-radius:4px;min-width:2px;}}
.hbar-inline{{font-size:11px;color:#475569;font-weight:600;margin-left:8px;white-space:nowrap;}}

/* Volume chart */
.hbar-row.hbar-vol{{grid-template-columns:95px 1fr;margin-bottom:10px;}}
.hbar-mono{{font-family:'SF Mono','Consolas','Courier New',monospace;font-size:11px;color:#475569;}}
.vol-xaxis{{display:flex;justify-content:space-between;margin-top:8px;padding-left:105px;font-size:10px;color:#94A3B8;}}
.vol-filter-label{{font-size:12px;font-weight:500;color:#2563EB;margin-left:8px;}}

/* Table */
.table-scroll{{max-height:2200px;overflow-y:auto;border-radius:10px;box-shadow:0 1px 4px rgba(0,0,0,.07);background:#fff;}}
table{{width:100%;border-collapse:collapse;font-size:12px;}}
th{{background:#F1F5F9;padding:8px 10px;text-align:left;font-weight:600;color:#475569;font-size:11px;border-bottom:2px solid #E2E8F0;position:sticky;top:0;z-index:10;white-space:nowrap;}}
td{{padding:7px 10px;border-bottom:1px solid #F1F5F9;vertical-align:middle;}}
tr:hover td{{background:#F8FAFC;}}
.dim{{color:#64748B;font-size:11px;}}
.th-sort,.th-filter{{color:#CBD5E1;cursor:pointer;font-size:11px;margin-left:2px;transition:color .15s;}}
.th-sort:hover,.th-filter:hover{{color:#2563EB;}}
.th-sort.active{{color:#2563EB;}}
.st-ok{{color:#16A34A;font-weight:600;}}
.st-unavailable{{color:#D97706;}}
.st-not_found,.st-error{{color:#DC2626;}}

/* Filter popup */
.filter-popup{{display:none;position:absolute;background:#fff;border:1px solid #E2E8F0;border-radius:10px;padding:14px;z-index:1000;min-width:210px;max-height:340px;overflow-y:auto;box-shadow:0 4px 20px rgba(0,0,0,.12);}}
.filter-popup.show{{display:block;}}
.fp-hint{{font-size:11px;color:#94A3B8;margin-bottom:8px;}}
.filter-popup label{{display:block;padding:3px 0;font-size:12px;color:#334155;cursor:pointer;}}
.filter-popup input[type="checkbox"]{{margin-right:6px;accent-color:#2563EB;}}
.filter-popup input[type="number"]{{width:80px;border:1px solid #E2E8F0;padding:4px 6px;border-radius:6px;font-size:12px;}}
.fp-actions{{margin-top:10px;display:flex;gap:8px;}}
.fp-btn{{padding:5px 14px;border-radius:6px;border:none;cursor:pointer;font-size:12px;font-weight:600;}}
.fp-apply{{background:#2563EB;color:#fff;}}
.fp-apply:hover{{background:#1D4ED8;}}
.fp-clear{{background:#F1F5F9;color:#475569;}}
.fp-clear:hover{{background:#E2E8F0;}}
.fp-selectall{{font-weight:700;padding-bottom:4px;border-bottom:1px solid #F1F5F9;margin-bottom:4px;}}
canvas{{width:100%!important;}}
</style>
</head>
<body>
<div class="header">
  <h1>{title}</h1>
  {'<div class="sub">' + subtitle + '</div>' if subtitle else ''}
</div>
<div class="container">
  <div class="tabs">{tabs_html}</div>
  {panels_html}
</div>
<div class="filter-popup" id="filterPopup"></div>

<script>
const rowData={row_json};
const volumeSeries={vol_json};
const COLORS=['#7C3AED','#F59E0B','#16A34A','#DC2626','#2563EB','#0891B2','#EA580C','#4F46E5','#059669','#DB2777','#7C3AED','#F59E0B','#16A34A','#DC2626','#2563EB'];
const TIER_COLORS=['#16A34A','#2563EB','#DC2626'];

// Compute price tiers per market
function computeTierBounds(mk){{
  const prices=rowData.filter(r=>r.market===mk&&r.price).map(r=>r.price);
  if(!prices.length)return[];
  const mn=Math.min(...prices),mx=Math.max(...prices),step=(mx-mn)/3;
  return[{{name:'Entry',lo:mn,hi:mn+step}},{{name:'Mid-tier',lo:mn+step,hi:mn+step*2}},{{name:'Premium',lo:mn+step*2,hi:mx+1}}];
}}
function getPriceTier(price,mk){{
  if(!price)return null;
  const tiers=computeTierBounds(mk);
  for(let i=0;i<tiers.length;i++){{if(price>=tiers[i].lo&&(i===2||price<tiers[i].hi))return tiers[i].name;}}
  return tiers.length?tiers[tiers.length-1].name:null;
}}

// Tabs
document.querySelectorAll('.tab').forEach(t=>{{
  t.addEventListener('click',()=>{{
    document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
    document.querySelectorAll('.tab-content').forEach(x=>{{x.classList.remove('active');x.style.display='none';}});
    t.classList.add('active');
    const panel=document.querySelector(`.tab-content[data-market="${{t.dataset.market}}"]`);
    if(panel){{panel.classList.add('active');panel.style.display='block';}}
  }});
}});

// === Volume chart (dynamic) ===
function renderVolChart(mk,filterType,filterValue){{
  const container=document.getElementById('vol-chart-'+mk);
  const label=document.getElementById('vol-filter-'+mk);
  if(!container)return;
  let items=rowData.filter(r=>r.market===mk&&r.search_volume);
  if(filterType==='price'){{
    items=items.filter(r=>getPriceTier(r.price,mk)===filterValue);
    if(label)label.textContent='[ '+filterValue+' ]';
  }}else if(filterType==='brand'){{
    items=items.filter(r=>r.brand===filterValue);
    if(label)label.textContent='[ '+filterValue+' ]';
  }}else{{
    if(label)label.textContent='';
  }}
  items.sort((a,b)=>(b.search_volume||0)-(a.search_volume||0));
  items=items.slice(0,15);
  if(!items.length){{container.innerHTML='<div style="color:#94A3B8;padding:20px;font-size:12px">No data for this filter</div>';return;}}
  const maxV=items[0].search_volume||1;
  let html='';
  items.forEach((r,i)=>{{
    const pct=(r.search_volume/maxV*100).toFixed(1);
    const c=COLORS[i%COLORS.length];
    html+=`<div class="hbar-row hbar-vol"><div class="hbar-label hbar-mono">${{r.asin}}</div><div class="hbar-track"><div class="hbar-fill" style="width:${{pct}}%;background:${{c}}"></div><span class="hbar-inline">${{r.brand||''}}</span></div></div>`;
  }});
  html+='<div class="vol-xaxis">';
  for(let i=0;i<=6;i++){{const v=maxV*i/6;html+=`<span>${{v>=1000?(v/1000).toFixed(1)+'k':Math.round(v)}}</span>`;}}
  html+='</div>';
  container.innerHTML=html;
}}
// Init volume charts
document.querySelectorAll('[id^="vol-chart-"]').forEach(el=>{{
  renderVolChart(el.id.replace('vol-chart-',''),null,null);
}});

// === Tier bar chart (canvas) ===
document.querySelectorAll('[id^="tier-bar-"]').forEach(canvas=>{{
  const mk=canvas.id.replace('tier-bar-','');
  const tiers=computeTierBounds(mk);
  const prices=rowData.filter(r=>r.market===mk&&r.price).map(r=>r.price);
  const counts=tiers.map(t=>prices.filter(p=>p>=t.lo&&(t.name==='Premium'||p<t.hi)).length);
  const maxC=Math.max(...counts)||1;
  const ctx=canvas.getContext('2d');
  const W=canvas.offsetWidth||400,H=220;
  canvas.width=W;canvas.height=H;
  const pad=40,barW=(W-2*pad)/(counts.length*2+1);
  // Y-axis
  ctx.strokeStyle='#E2E8F0';ctx.lineWidth=0.5;
  for(let i=0;i<=4;i++){{
    const y=pad+(H-pad-30)*(1-i/4);
    ctx.beginPath();ctx.moveTo(pad,y);ctx.lineTo(W-10,y);ctx.stroke();
    ctx.fillStyle='#94A3B8';ctx.font='10px Segoe UI';
    ctx.fillText(Math.round(maxC*i/4),5,y+3);
  }}
  // Bars
  counts.forEach((c,i)=>{{
    const x=pad+barW*(i*2+1);
    const barH=(H-pad-30)*c/maxC;
    const y=H-30-barH;
    ctx.fillStyle=TIER_COLORS[i];
    ctx.beginPath();ctx.roundRect(x,y,barW,barH,[4,4,0,0]);ctx.fill();
    // Count label
    ctx.fillStyle='#1E293B';ctx.font='bold 13px Segoe UI';ctx.textAlign='center';
    ctx.fillText(c,x+barW/2,y-6);
    // X label
    ctx.fillStyle='#475569';ctx.font='11px Segoe UI';
    const lbl=tiers[i].name==='Entry'?'$0\u2013$'+Math.round(tiers[i].hi):tiers[i].name==='Mid-tier'?'$'+Math.round(tiers[i].lo)+'\u2013$'+Math.round(tiers[i].hi):'$'+Math.round(tiers[i].lo)+'+';
    ctx.fillText(lbl,x+barW/2,H-10);
    ctx.textAlign='left';
  }});
}});

// === Click interaction (price/brand -> filter volume) ===
document.querySelectorAll('.clickable-bar').forEach(bar=>{{
  bar.addEventListener('click',()=>{{
    const mk=bar.dataset.market,type=bar.dataset.filterType,val=bar.dataset.filterValue;
    const wasSelected=bar.classList.contains('selected');
    // Clear all selections in same market
    document.querySelectorAll(`.clickable-bar[data-market="${{mk}}"]`).forEach(b=>b.classList.remove('selected'));
    if(!wasSelected){{
      bar.classList.add('selected');
      renderVolChart(mk,type,val);
    }}else{{
      renderVolChart(mk,null,null);
    }}
  }});
}});

// === Table sort ===
document.querySelectorAll('.th-sort').forEach(el=>{{
  el.addEventListener('click',e=>{{
    e.stopPropagation();
    const th=el.closest('th'),table=th.closest('table'),tbody=table.querySelector('tbody');
    const col=parseInt(el.dataset.col),rows=Array.from(tbody.querySelectorAll('tr'));
    const wasAsc=el.classList.contains('asc');
    table.querySelectorAll('.th-sort').forEach(s=>{{s.classList.remove('asc','desc','active');s.innerHTML='&#8597;';}});
    const asc=!wasAsc;
    el.classList.add(asc?'asc':'desc','active');
    el.innerHTML=asc?'&#9650;':'&#9660;';
    rows.sort((a,b)=>{{
      let va=a.cells[col]?.textContent.trim()||'',vb=b.cells[col]?.textContent.trim()||'';
      const na=parseFloat(va.replace(/[^\\d.\\-]/g,'')),nb=parseFloat(vb.replace(/[^\\d.\\-]/g,''));
      if(!isNaN(na)&&!isNaN(nb))return asc?na-nb:nb-na;
      return asc?va.localeCompare(vb):vb.localeCompare(va);
    }});
    rows.forEach(r=>tbody.appendChild(r));
  }});
}});

// === Table filter ===
const popup=document.getElementById('filterPopup');
let fCol=null,fTable=null;
document.querySelectorAll('.th-filter').forEach(el=>{{
  el.addEventListener('click',e=>{{
    e.stopPropagation();
    const th=el.closest('th'),table=th.closest('table'),col=parseInt(el.dataset.col);
    fCol=col;fTable=table;
    const rows=Array.from(table.querySelector('tbody').querySelectorAll('tr'));
    const vals=new Set();let isNum=true;
    rows.forEach(r=>{{const v=r.cells[col]?.textContent.trim()||'';vals.add(v);if(v&&v!=='-'&&isNaN(parseFloat(v.replace(/[^\\d.\\-]/g,''))))isNum=false;}});
    let h='';
    if(isNum&&vals.size>5){{
      h=`<div class="fp-hint">Numeric range</div><label>Min: <input type="number" id="fMin" step="any"></label><label style="margin-top:4px">Max: <input type="number" id="fMax" step="any"></label><div class="fp-actions"><button class="fp-btn fp-apply" onclick="applyNum()">Apply</button><button class="fp-btn fp-clear" onclick="clearF()">Clear</button></div>`;
    }}else{{
      const sorted=Array.from(vals).sort();
      h=`<div class="fp-hint">Select values</div><label class="fp-selectall"><input type="checkbox" id="fpSelectAll" checked onchange="toggleAll(this)"> <strong>Select All</strong></label>`;
      sorted.forEach(v=>{{h+=`<label><input type="checkbox" class="fp-item-cb" value="${{v.replace(/"/g,'&quot;')}}" checked> ${{v||'(empty)'}}</label>`;}});
      h+=`<div class="fp-actions"><button class="fp-btn fp-apply" onclick="applyTxt()">Apply</button><button class="fp-btn fp-clear" onclick="clearF()">Clear</button></div>`;
    }}
    popup.innerHTML=h;
    const rect=el.getBoundingClientRect();
    popup.style.top=(rect.bottom+window.scrollY+4)+'px';
    popup.style.left=Math.min(rect.left+window.scrollX,window.innerWidth-240)+'px';
    popup.classList.add('show');
  }});
}});
document.addEventListener('click',e=>{{if(!popup.contains(e.target)&&!e.target.classList.contains('th-filter'))popup.classList.remove('show');}});
function toggleAll(el){{popup.querySelectorAll('.fp-item-cb').forEach(cb=>cb.checked=el.checked);}}
function applyNum(){{
  const mn=parseFloat(document.getElementById('fMin').value),mx=parseFloat(document.getElementById('fMax').value);
  Array.from(fTable.querySelector('tbody').rows).forEach(r=>{{
    const v=parseFloat((r.cells[fCol]?.textContent||'').replace(/[^\\d.\\-]/g,''));
    let ok=true;if(!isNaN(mn)&&(isNaN(v)||v<mn))ok=false;if(!isNaN(mx)&&(isNaN(v)||v>mx))ok=false;
    r.style.display=ok?'':'none';
  }});
  popup.classList.remove('show');
}}
function applyTxt(){{
  const ck=new Set();popup.querySelectorAll('.fp-item-cb:checked').forEach(c=>ck.add(c.value));
  Array.from(fTable.querySelector('tbody').rows).forEach(r=>{{
    r.style.display=ck.has(r.cells[fCol]?.textContent.trim()||'')?'':'none';
  }});
  popup.classList.remove('show');
}}
function clearF(){{Array.from(fTable.querySelector('tbody').rows).forEach(r=>r.style.display='');popup.classList.remove('show');}}
</script>
</body>
</html>'''
